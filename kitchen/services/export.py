from io import BytesIO

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

BRAND = '1A5542'
BRAND_LIGHT = 'D4EBE2'
ACCENT = 'B56A28'
INK = '12201A'
MUTED = '5A6F63'
WHITE = 'FFFFFF'
ROW_ALT = 'F4F8F5'

THIN = Side(style='thin', color='C5D4CB')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill('solid', fgColor=BRAND)
HEADER_FONT = Font(name='Calibri', bold=True, color=WHITE, size=11)
TITLE_FONT = Font(name='Calibri', bold=True, color=INK, size=16)
SUB_FONT = Font(name='Calibri', color=MUTED, size=10)
META_FONT = Font(name='Calibri', color=MUTED, size=9, italic=True)
BODY_FONT = Font(name='Calibri', color=INK, size=10)
TOTAL_FILL = PatternFill('solid', fgColor=BRAND_LIGHT)
TOTAL_FONT = Font(name='Calibri', bold=True, color=INK, size=10)
ALT_FILL = PatternFill('solid', fgColor=ROW_ALT)


def _money(value):
    try:
        n = float(value or 0)
    except (TypeError, ValueError):
        return '0'
    return f'{n:,.0f}'.replace(',', ' ')


def _now_label():
    return timezone.localtime().strftime('%d.%m.%Y %H:%M')


def csv_response(filename, title, subtitle, headers, rows, totals=None):
    lines = [
        f'"{title}"',
        f'"{subtitle or ""}"',
        f'"Yaratilgan: {_now_label()}"',
        '""',
        ';'.join(f'"{h}"' for h in headers),
    ]
    for row in rows:
        cells = []
        for cell in row:
            text = str(cell).replace('"', '""')
            cells.append(f'"{text}"')
        lines.append(';'.join(cells))
    if totals:
        lines.append('')
        lines.append(';'.join(f'"{c}"' for c in totals))
    payload = '\ufeff' + '\n'.join(lines)
    response = HttpResponse(payload, content_type='text/csv; charset=utf-8')
    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
    return response


def excel_response(filename, title, subtitle, headers, rows, totals=None, numeric_cols=None):
    numeric_cols = numeric_cols or set()
    wb = Workbook()
    ws = wb.active
    ws.title = 'Hisobot'

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(len(headers), 4))
    c1 = ws.cell(row=1, column=1, value=title)
    c1.font = TITLE_FONT
    c1.alignment = Alignment(vertical='center')

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(len(headers), 4))
    c2 = ws.cell(row=2, column=1, value=subtitle or '')
    c2.font = SUB_FONT

    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max(len(headers), 4))
    c3 = ws.cell(row=3, column=1, value=f'Yaratilgan: {_now_label()}')
    c3.font = META_FONT

    start = 5
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=start, column=col, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal='center', vertical='center')

    for r_idx, row in enumerate(rows):
        row_num = start + 1 + r_idx
        fill = ALT_FILL if r_idx % 2 else PatternFill('solid', fgColor=WHITE)
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_num, column=c_idx, value=value)
            cell.font = BODY_FONT
            cell.fill = fill
            cell.border = BORDER
            if (c_idx - 1) in numeric_cols:
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if isinstance(value, (int, float)) or (
                    isinstance(value, str) and value.replace('.', '', 1).replace('-', '', 1).isdigit()
                ):
                    try:
                        cell.value = float(value)
                        cell.number_format = '#,##0'
                    except (TypeError, ValueError):
                        pass
            else:
                cell.alignment = Alignment(vertical='center', wrap_text=True)

    if totals:
        total_row = start + 1 + len(rows) + 1
        for c_idx, value in enumerate(totals, 1):
            cell = ws.cell(row=total_row, column=c_idx, value=value)
            cell.font = TOTAL_FONT
            cell.fill = TOTAL_FILL
            cell.border = BORDER
            if c_idx > 1:
                cell.alignment = Alignment(horizontal='right', vertical='center')

    for col in range(1, len(headers) + 1):
        letter = get_column_letter(col)
        max_len = len(str(headers[col - 1]))
        for row in rows:
            if col - 1 < len(row):
                max_len = max(max_len, len(str(row[col - 1])))
        if totals and col - 1 < len(totals):
            max_len = max(max_len, len(str(totals[col - 1])))
        ws.column_dimensions[letter].width = min(max(max_len + 3, 12), 42)

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[start].height = 22
    ws.freeze_panes = ws.cell(row=start + 1, column=1)

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
    return response


def spreadsheet_download(request, filename, title, subtitle, headers, rows, totals=None, numeric_cols=None):
    fmt = (request.GET.get('format') or 'xlsx').lower()
    if fmt == 'csv':
        return csv_response(filename, title, subtitle, headers, rows, totals=totals)
    return excel_response(filename, title, subtitle, headers, rows, totals=totals, numeric_cols=numeric_cols)
